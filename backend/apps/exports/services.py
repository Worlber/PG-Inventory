import csv
import io

from openpyxl import Workbook

from apps.inventory.models import DatabaseInfo, DatabaseUser, PostgreSQLInstance


class ExportService:
    def export_databases_csv(self, instance_id):
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(["Database Name", "Size (Bytes)", "Size (Human)", "Last Seen"])

        for db in DatabaseInfo.objects.filter(instance_id=instance_id).order_by("-size_bytes"):
            writer.writerow([db.name, db.size_bytes, self._human_size(db.size_bytes), db.last_seen])

        output.seek(0)
        return output

    def export_users_csv(self, instance_id):
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(["Username", "Superuser", "Can Login", "Permissions", "Last Seen"])

        for user in DatabaseUser.objects.filter(instance_id=instance_id).order_by("username"):
            writer.writerow([
                user.username,
                user.is_superuser,
                user.can_login,
                ", ".join(user.permissions) if user.permissions else "",
                user.last_seen,
            ])

        output.seek(0)
        return output

    def export_full_inventory_csv(self):
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow([
            "Hostname", "IP Address", "Port", "Environment",
            "Application", "HA Group", "Status", "Role",
            "PG Version", "OS Version", "RAM (MB)", "CPU Count",
            "Database Count", "Last Checked",
        ])

        for inst in PostgreSQLInstance.objects.select_related("application", "ha_group").all():
            writer.writerow([
                inst.hostname,
                inst.ip_address,
                inst.port,
                inst.get_environment_display(),
                inst.application.name if inst.application else "",
                inst.ha_group.name if inst.ha_group else "",
                "UP" if inst.is_up else "DOWN",
                inst.role,
                inst.pg_version,
                inst.os_version,
                inst.ram_mb or "",
                inst.cpu_count or "",
                inst.databases.count(),
                str(inst.last_checked) if inst.last_checked else "",
            ])

        output.seek(0)
        return output

    def export_full_inventory_xlsx(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Inventory"

        headers = [
            "Hostname", "IP Address", "Port", "Environment",
            "Application", "HA Group", "Status", "Role",
            "PG Version", "OS Version", "RAM (MB)", "CPU Count",
            "Database Count", "Last Checked",
        ]
        ws.append(headers)

        for inst in PostgreSQLInstance.objects.select_related("application", "ha_group").all():
            ws.append([
                inst.hostname,
                inst.ip_address,
                inst.port,
                inst.get_environment_display(),
                inst.application.name if inst.application else "",
                inst.ha_group.name if inst.ha_group else "",
                "UP" if inst.is_up else "DOWN",
                inst.role,
                inst.pg_version,
                inst.os_version,
                inst.ram_mb or "",
                inst.cpu_count or "",
                inst.databases.count(),
                str(inst.last_checked) if inst.last_checked else "",
            ])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def export_databases_xlsx(self, instance_id):
        wb = Workbook()
        ws = wb.active
        ws.title = "Databases"
        ws.append(["Database Name", "Size (Bytes)", "Size (Human)", "Last Seen"])

        for db in DatabaseInfo.objects.filter(instance_id=instance_id).order_by("-size_bytes"):
            ws.append([db.name, db.size_bytes, self._human_size(db.size_bytes), str(db.last_seen)])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def export_users_xlsx(self, instance_id):
        wb = Workbook()
        ws = wb.active
        ws.title = "Users"
        ws.append(["Username", "Superuser", "Can Login", "Permissions", "Last Seen"])

        for user in DatabaseUser.objects.filter(instance_id=instance_id).order_by("username"):
            ws.append([
                user.username,
                user.is_superuser,
                user.can_login,
                ", ".join(user.permissions) if user.permissions else "",
                str(user.last_seen),
            ])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    @staticmethod
    def _human_size(size_bytes):
        size = float(size_bytes)
        for unit in ["B", "KB", "MB", "GB", "TB"]:
            if size < 1024:
                return f"{size:.1f} {unit}"
            size /= 1024
        return f"{size:.1f} PB"
